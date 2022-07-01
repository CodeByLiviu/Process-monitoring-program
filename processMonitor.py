'''
THE PROBLEM:
Implement a program that will launch a specified process and periodically (with a provided time interval) collect the following data about it:
•	CPU usage (percent);
•	Memory consumption: Working Set and Private Bytes (for Windows systems) or Resident Set Size and Virtual Memory Size (for Linux systems);
•	Number of open handles (for Windows systems) or file descriptors (for Linux systems).
Data collection should be performed all the time the process is running.
Path to the executable file for the process and time interval between data collection iterations should be provided by user.
Collected data should be stored on the disk. Format of stored data should support automated parsing to potentially allow, for example, drawing of charts.
'''


'''
                                          READ THIS
                        ***   Steps to fallow to use this program   ***
1. Choose the folder where the excel file will be stored!
2. Create an excel file in the chosen folder and name it, the excel file must end with the extension .elsx, for example: RaportStatus.elsx.
3. Open the excel file and write what is in model: 9 cells, this is the first and only line we wrote horizontally (row arranged on horizontal). 
   A row arranges data horizontally from left to right. 
   model: 1. Time,	2. PID,	3. Process CPU utilization (%),	4. Process Memory utilization,	5. Process Memory utilization(%),	6. Number of open handles,	7. Number of open threads,	8. Memory usage on the whole system (%),	9. Cpu usage on the whole system (%);
4. Save the excel file, exit, and copy the path to the excel file!
5. Choose the interval(seconds) between updates, it must be a integer number!
6. Choose the PID, you can find it in taskmanager (windows system), in section "Details", PID must be an integer!
6. Start the program enter and enter these values: interval = 5, pid = 6060, path = C:/Users/Liviu/Desktop/data/StatusReport.xlsx 
THIS PROGRAM STORES THE DESIRED DATA FOR A PROCESS THAT IS ALREADY RUNNING, OR WILL BE STARTED, YOU HAVE TO GIVE IT THE PID(MANUALLY) AND PUT IT IN THIS PROGRAM TO SEE THE PROCESS STATISTICS!
                              THIS PROGRAM DOES NOT START A DESIRED PROCESS!
'''
# when you will use the program make sure the excel file is closed.
# if it's not: you will get some errors!
# you can't use the program while excel file is open!
# some PIDs can't be used, because they have some windows protection.
# Example of PIDs with protection: SYSTEM, LOCAL SERVICE, NETWORK SERVICE, DWM-2, UMFD-0...
# If you want to try one of this pids with protection, you will get an error "Acces denied"!
# try another PID (from an app)
# make sure you use this "/" and not this "\" when entering the file path.
# after every update, all the data is stored in excel file.
# THIS PROGRAM DOES NOT HAVE A STOP BUTTON, YOU HAVE TO CLOSE IT!


# The solution:
# import the relevant modules
import psutil
import datetime
import time
import openpyxl


def getMemUsage():
      '''
      This method is used to get the memory usage overall
      :return: returns a value (the percentage of memory usage on the whole system)
      '''
      ramUsage = psutil.virtual_memory().percent
      return ramUsage


def getCpuUsage():
      '''
      This method is used to get the cpu usage overall
      :return: returns a value (the percentage of CPU usage on the whole system)
      '''
      cpuUsage = psutil.cpu_percent(interval=1)
      return cpuUsage


def getProcessStatus(pid, path):
      '''
      This method is used to get all the data we want and store it in aa excel file.
      :param pid: is the path to a process ID
      :param path: the to the .excel file
      :return: returns a bool that can stop the process if there is something wrong.
      '''
      try:
            pid = int(pid)
            selectedProcess = psutil.Process(pid)
      except:
            print("The PID is not working...")
            return False
      time = datetime.datetime.now().strftime("%Y:%m:%d - %H:%M:%S")
      cpu = selectedProcess.cpu_percent(interval = 1)/psutil.cpu_count()
      memoryMb = selectedProcess.memory_full_info().rss/(1024 * 1024)
      memory = selectedProcess.memory_percent()
      numberOfOpenHandles = selectedProcess.num_handles()
      numberOfOpenThreads = selectedProcess.num_threads()
      try:
            file = openpyxl.load_workbook(path)
      except:
            print("The Path is not working...")
            return False
      sheet = file.active
      sheet.cell(column = 1, row = sheet.max_row + 1, value = time)
      sheet.cell(column = 2, row = sheet.max_row, value = pid)
      sheet.cell(column = 3, row = sheet.max_row, value = cpu)
      sheet.cell(column = 4, row = sheet.max_row, value = memoryMb)
      sheet.cell(column = 5, row = sheet.max_row, value = memory)
      sheet.cell(column = 6, row = sheet.max_row, value = numberOfOpenHandles)
      sheet.cell(column = 8, row = sheet.max_row, value = numberOfOpenThreads)
      sheet.cell(column = 7, row = sheet.max_row, value = getCpuUsage())
      sheet.cell(column = 9, row = sheet.max_row, value = getMemUsage())
      file.save(path)


def mainFunction(interval, pid, path):
      '''
      This method is used to combine all the functions to get the desired result (an excel file with dates)
      :param interval: time in seconds between each update
      :param pid: is the path to the process ID
      :param path: the to the .excel file
      :return: it is used stop the loop
      '''
      while True:
            try:
                  interval = int(interval)
            except:
                  print("The interval is not an integer number, please try again!")
                  return
            status = getProcessStatus(pid, path)
            if status == False:
                  print("The program will stop!")
                  return
            dialog = "The process is running... all the data is writing in the excel file now!"
            print(dialog)
            time.sleep(interval)
            getMemUsage()
            getCpuUsage()
            getProcessStatus(pid, path)


# the path to the excel folder should look like this:
# be sure you will use this '/' not this '\'.
# path = "C:/Users/Liviu/Desktop/data/StatusReport.xlsx"

interval = input("Enter the interval between updates: ")
pid = input("Enter PID ID: ")
path = input("Enter the location of the excel file: ")

mainFunction(interval, pid, path)